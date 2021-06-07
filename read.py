from openpyxl import load_workbook, Workbook
wb = load_workbook('contacts.xlsx')
sheet = wb.active
print(sheet)
max_column = sheet.max_column
max_row = sheet.max_row
print('max row = ', max_row)
print('max column = ', max_column)


for item in range(1, max_row+1):
    print('row = ', item, 'mobile number = ', sheet.cell(item, 1).value)
for name in range(1, max_row+1):
    print('row = ', name, 'name = ', sheet.cell(name, 2).value)

for item in range(1, max_row+1):
    print('mobile number = ', sheet.cell(item, 1).value, "name = ", sheet.cell(item, 2).value)
